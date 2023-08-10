import math
import string
import random
import numpy as np
import openpyxl
import os
import pandas as pd
import collections

class ExcelPricing:

    def __init__(self, excel_file, P0=1000):

        self.excel_file = excel_file

        self.P0 = P0

    def get_price(self, matrix=None):


        # 加载工作簿
        wb = openpyxl.load_workbook(self.excel_file)
        worksheet = wb.active

        # 获取行数、列数
        n = worksheet.max_row
        m = worksheet.max_column

        # 计算单元格平均宽高
        w = self.get_avg_col_width(worksheet, m)
        h = self.get_avg_row_height(worksheet, n)

        # 读取单元格的值

        matrix = self.get_matrix(worksheet, n, m)
        self.matrix = matrix
        print(matrix)
        # 计算信息量
        total_info = self.calc_info_default(n, m, matrix)

        # 获取文件大小
        file_size = os.path.getsize(self.excel_file)
        volume = n * m * w * h
        # 计算价格

        density = total_info / volume
        price = self.P0 * density

        return price

    def get_avg_col_width(self, worksheet, m):
        # 计算单元格平均宽
        w = 0
        for i in range(1, m + 1):
            if i <= 26:
                col = string.ascii_uppercase[i - 1]
            else:
                # 超过26时需要根据i计算具体列名
                col = string.ascii_uppercase[(i - 1) % 26] + string.ascii_uppercase[(i - 1) // 26 - 1]

            w += worksheet.column_dimensions[col].width

        w = w / m
        return w

    def get_avg_row_height(self, worksheet, n):
        # 计算单元格平均高
        h = 0
        for j in range(1, n + 1):
            h += worksheet.row_dimensions[j].height

        h = h / n
        return h

    def get_matrix(self, worksheet, n, m):

        # 读取每个单元格的值
        matrix = []
        for i in range(1, n + 1):
            row = []
            for j in range(1, m + 1):
                cell = worksheet.cell(row=i, column=j)
                value = cell.value
                row.append(value)
            matrix.append(row)

        return matrix

    def calc_info_default(self, n, m, matrix):
        # 计算信息量
        total_info = 0
        for i in range(n):
            for j in range(m):
                # 假设每个单元格值出现频率相同
                p = 1 / (n * m)
                info = -math.log2(p)
                total_info += info
        # print(total_info)
        return total_info


# 使用示例:

pricing = ExcelPricing('lib/有色金属测试数据.xlsx')
price = pricing.get_price()
print(price)
