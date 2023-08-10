import os
import pandas as pd
import numpy as np
import math
import openpyxl


class DynamicExcelPricing:
    def __init__(self, excel_file, alpha=1.0, beta=10):
        self.excel_file = excel_file
        self.alpha = alpha
        self.beta = beta
        # 初始化表格数据
        wb = openpyxl.load_workbook(excel_file)
        worksheet = wb.active
        self.n = worksheet.max_row
        self.m = worksheet.max_column
        self.matrix = self.get_matrix(worksheet)
        self.file_size = os.path.getsize(excel_file) / 1024
        self.base_size = self.file_size / (self.m * self.n)  # 单元格平均大小
        # print("filesize", self.file_size)
        # print("basesize:", self.base_size)
        # 当前单元格数量
        self.cell_count = self.n * self.m
        self.threshold=25000

    def get_matrix(self, worksheet):
        # 从Excel文件中读取初始矩阵数据
        matrix = []
        for i in range(1, self.n + 1):
            row = []
            for j in range(1, self.m + 1):
                cell = worksheet.cell(row=i, column=j)
                value = cell.value
                row.append(value)
            matrix.append(row)
        return matrix

    def update_matrix(self):
        # 模拟表格数据更新
        new_rows = np.random.randint(low=1, high=100)
        new_cols = np.random.randint(low=1, high=100)
        new_data = pd.DataFrame(np.random.rand(new_rows, new_cols))

        # 使用pandas拼接新数据
        self.matrix = pd.concat([pd.DataFrame(self.matrix), new_data], axis=0)
        self.n = self.matrix.shape[0]
        self.m = self.matrix.shape[1]

        # 更新单元格数量
        if self.cell_count==0:
            self.threshold = np.random.randint(low=5000, high=100000)  # 阈值，即单元格数量大于该值时开始增加文件大小


        self.cell_count += new_rows * new_cols

        # 文件增长，当增加的单元格数大于阈值时，增加文件大小

        if self.cell_count > self.threshold:
            self.file_size += ((self.cell_count - self.threshold)) * self.base_size
            self.cell_count = 0
        # print("cellcount:",self.cell_count)
        # print("filesize", self.file_size)

    def get_price(self):
        # 计算价格
        total_info = self.calc_info()
        volume = self.n * self.m  # 行数x列数=体积
        density = total_info / volume  # 信息密度，信息量除以体积
        print(volume,density)
        # 使用回归方程将文件大小映射到价格范围(0, 10000)
        size = self.file_size
        price = 1130.15 + 11.07 * (size ** 1) + 0.25 * (size ** 2)

        # 使用sigmoid函数将信息密度映射到价格范围(0, 10000)
        price = price * (1 - 1 / (1 + np.exp(-self.alpha * density)))/self.beta
        return price

    def calc_info(self):
        # 计算信息量
        total_info = 0
        for i in range(self.n):
            for j in range(self.m):
                p = 1 / (self.n * self.m)  # 假设每个单元格的概率相同
                info = -math.log2(p)  # 使用信息熵公式计算信息量
                total_info += info
        print("totalinfo:",total_info)
        return total_info


# 使用示例
# pricing = DynamicExcelPricing('有色金属测试数据.xlsx', alpha=0.3)  # 设置alpha值，调节价格对信息密度的敏感度
# 创建一个DataFrame用于存储输出结果
# output_data = pd.DataFrame(columns=['Price', 'FileSize', 'TotalInfo'])
#
# for i in range(50):
#     pricing.update_matrix()
#     print("Price:", pricing.get_price())
#     price = pricing.get_price()
#     filesize = pricing.file_size
#     totalinfo = pricing.calc_info()
#
#     # 将当前结果添加到DataFrame
#     output_data = output_data.append({'Price': price, 'FileSize': filesize, 'TotalInfo': totalinfo}, ignore_index=True)
#
# # 将DataFrame保存到Excel文件中
# output_file = 'output.xlsx'
# output_data.to_excel(output_file, index=False)
# int("输出已保存到文件:"+ output_file)