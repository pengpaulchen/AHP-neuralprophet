import os
import pandas as pd
import numpy as np
import math
import random
import openpyxl


class DynamicExcelPricing:
    def __init__(self, excel_file, P0=100):
        self.excel_file = excel_file
        self.P0 = P0

        # 初始化表格数据
        wb = openpyxl.load_workbook(excel_file)
        worksheet = wb.active
        self.n = worksheet.max_row
        self.m = worksheet.max_column
        self.matrix = self.get_matrix(worksheet)
        self.file_size = os.path.getsize(excel_file) / 1024
        self.base_size = self.file_size / (self.m * self.n)  # 单元格平均大小
        print("filesize", self.file_size)
        print("basesize:", self.base_size)
        # 当前单元格数量
        self.cell_count = self.n * self.m

    def get_matrix(self, worksheet):
        # 从Excel文件中读取初始矩阵数据
        matrix = []
        for i in range(1, self.n + 1):
            row = []
            for j in range(1, self.m + 1):
                cell = worksheet.cell(row=i, column=j)
                value = cell.value
                row.append(value)
            matrix.append(row)
        return matrix

    def update_matrix(self):
        # 模拟表格数据更新
        new_rows = np.random.randint(low=1, high=5)
        new_cols = np.random.randint(low=1, high=5)
        new_data = pd.DataFrame(np.random.rand(new_rows, new_cols))

        # 使用pandas拼接新数据
        self.matrix = pd.concat([pd.DataFrame(self.matrix), new_data], axis=0)
        self.n = self.matrix.shape[0]
        self.m = self.matrix.shape[1]
        # 更新单元格数量
        self.cell_count += new_rows * new_cols

    def get_price(self):
        # 计算价格
        total_info = self.calc_info()
        volume = self.n * self.m  # 行数x列数=体积
        density = total_info / volume  # 信息密度，信息量除以体积
        price = self.P0*density
        # # 根据文件大小设定不同的价格范围
        # if self.file_size < 1024:  # 文件小于1GB
        #     price_range = (0, 100)
        # elif 1024 <= self.file_size < 10240:  # 文件在1GB到10GB之间
        #     price_range = (100, 500)
        # else:  # 文件大于等于10MB
        #     price_range = (500, 1000)
        #
        # # 将信息密度映射到相应的价格范围
        # price = np.interp(density, (0, 1), price_range)
        return price

    def calc_info(self):
        # 计算信息量
        total_info = 0
        for i in range(self.n):
            for j in range(self.m):
                p = 1 / (self.n * self.m)  # 假设每个单元格的概率相同
                info = -math.log2(p)  # 使用信息熵公式计算信息量
                total_info += info
        return total_info


# 使用示例
pricing = DynamicExcelPricing('lib/有色金属测试数据.xlsx')

for i in range(50):
    pricing.update_matrix()
    print("Price:", pricing.get_price())
